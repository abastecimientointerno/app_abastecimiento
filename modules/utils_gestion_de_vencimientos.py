import pandas as pd
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import ColorScaleRule

def process_data(mb52_df, mb25_df):
    # Establecer un formato estándar para las fechas
    mb52_df['Cad./FPC'] = pd.to_datetime(mb52_df['Cad./FPC'], format='%Y-%m-%d', errors='coerce')
    mb25_df['Fecha de necesidad'] = pd.to_datetime(mb25_df['Fecha de necesidad'], format='%Y-%m-%d', errors='coerce')
    
    # Definir un rango de fechas para filtrar
    fecha_limite = datetime.now() + timedelta(days=90)
    proximos_a_vencer = mb52_df[(mb52_df['Cad./FPC'] > datetime.now()) & (mb52_df['Cad./FPC'] <= fecha_limite)]
    
    resultados = []
    detalles = []
    
    for index, material in proximos_a_vencer.iterrows():
        centro = material['Centro']
        codigo_material = material['Material']
        descripcion = material['Texto breve de material']
        cantidad_disponible = material['Libre utilización']
        lote = material['Lote']
        valorizado = material['Valor libre util.']
        almacen = material['Almacén']
        ubicacion = material['Ubicación']
        fecha_vencimiento = material['Cad./FPC']

        # Filtrar reservas para el mismo material en el mismo centro y otros centros
        reservas_mismo_centro = mb25_df[(mb25_df['Texto breve de material'] == descripcion) & 
                                        (mb25_df['Centro'] == centro)]
        reservas_otro_centro = mb25_df[(mb25_df['Texto breve de material'] == descripcion) & 
                                        (mb25_df['Centro'] != centro)]
        
        # Calcular la cantidad reservada según el centro correspondiente
        if not reservas_mismo_centro.empty:
            estado = "Gestion"
            cantidad_reservada = reservas_mismo_centro['Cantidad diferencia'].sum()
            cantidad_a_utilizar = min(cantidad_disponible, cantidad_reservada)
            num_reservas = reservas_mismo_centro['Nº reserva'].tolist()
            detalles.append({
                'centro': centro,
                'codigo_material': codigo_material,
                'descripcion': descripcion,
                'cantidad': cantidad_a_utilizar,
                'valorizado' : valorizado,
                'lote': lote,
                'um': material['Unidad medida base'],
                'centro_necesidad': centro,
                'cantidad_reservada': reservas_mismo_centro['Cantidad diferencia'].tolist(),
                'reserva': num_reservas,
                'posicion': reservas_mismo_centro['Nº pos.reserva traslado'].tolist(),
                'estado': estado
            })
        elif not reservas_otro_centro.empty:
            estado = "Traslado"
            cantidad_reservada = reservas_otro_centro['Cantidad diferencia'].sum()
            cantidad_a_utilizar = min(cantidad_disponible, cantidad_reservada)
            num_reservas = reservas_otro_centro['Nº reserva'].tolist()
            centro_necesidad = reservas_otro_centro['Centro'].unique().tolist()
            detalles.append({
                'centro': centro,
                'codigo_material': codigo_material,
                'descripcion': descripcion,
                'cantidad': cantidad_a_utilizar,
                'valorizado' : valorizado,
                'lote': lote,
                'um': material['Unidad medida base'],
                'centro_necesidad': centro_necesidad,
                'cantidad_reservada': reservas_otro_centro['Cantidad diferencia'].tolist(),
                'reserva': num_reservas,
                'posicion': reservas_otro_centro['Nº pos.reserva traslado'].tolist(),
                'estado': estado
            })
        else:
            estado = "Sin necesidad"
            cantidad_reservada = 0
            num_reservas = []

        resultados.append({
            'centro': centro,
            'codigo_material': codigo_material,
            'descripcion': descripcion,
            'almacen': almacen,
            'lote': lote,
            'valorizado': valorizado,
            'ubicacion': ubicacion,
            'fecha_vencimiento': fecha_vencimiento,
            'cantidad': cantidad_disponible,  # Conservamos la cantidad total disponible en el reporte final
            'cantidad_reservada': cantidad_reservada,
            'estado': estado
        })
    
    resultados_df = pd.DataFrame(resultados)
    detalles_df = pd.DataFrame(detalles)
    
    # Asegurarnos de que las fechas están correctamente formateadas
    resultados_df['fecha_vencimiento'] = pd.to_datetime(resultados_df['fecha_vencimiento'], errors='coerce')
    
    return resultados_df, detalles_df



def create_excel(resultados_df, detalles_df):
    # Crear tabla resumen por centro y estado
    resumen_centro_estado = resultados_df.pivot_table(
        values='valorizado',
        index='centro',
        columns='estado',
        aggfunc='sum',
        fill_value=0
    ).reset_index()

    for estado in ['Gestion', 'Traslado', 'Sin necesidad']:
        if estado not in resumen_centro_estado.columns:
            resumen_centro_estado[estado] = 0
    
    resumen_centro_estado = resumen_centro_estado[['centro', 'Gestion', 'Traslado', 'Sin necesidad']]
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        resultados_df.to_excel(writer, sheet_name='Resumen', index=False)
        detalles_df.to_excel(writer, sheet_name='Detalle', index=False)
        resumen_centro_estado.to_excel(writer, sheet_name='Resumen por Centro', index=False)
        
        workbook = writer.book
        worksheet = workbook['Resumen por Centro']
        
        # Formatear celdas
        for col in ['B', 'C', 'D']:
            for cell in worksheet[col][1:]:
                cell.number_format = '"S/. "#,##0.00'
        
        # Normalización por fila y gráficos
        for row in worksheet.iter_rows(min_row=2, min_col=2, max_col=4):
            total = sum(cell.value for cell in row)
            for cell in row:
                cell.value = cell.value / total if total > 0 else 0
        
        for col in ['B', 'C', 'D']:
            for cell in worksheet[col][1:]:
                cell.number_format = '0.00%'
        
        # Crear gráfico de barras
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "percentStacked"
        chart.overlap = 100
        data = Reference(worksheet, min_col=2, min_row=1, max_row=worksheet.max_row, max_col=4)
        cats = Reference(worksheet, min_col=1, min_row=2, max_row=worksheet.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.title = "Estados por centro"
        worksheet.add_chart(chart, "F2")
        
        header = worksheet[1]
        for cell in header:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        worksheet.conditional_formatting.add(f'B2:D{worksheet.max_row}',
            ColorScaleRule(start_type='min', start_color='FFFFFF',
                           mid_type='percentile', mid_value=50, mid_color='FFFF00',
                           end_type='max', end_color='FF0000'))
    
    output.seek(0)
    return output.getvalue()